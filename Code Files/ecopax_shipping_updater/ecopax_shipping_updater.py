import multiprocessing
import threading
from ShippingUpdaterUtility import *
from ShippingUpdaterGUI import *
from ShippingContainerDB import *
from CMA import cma_search
from Cosco import cosco_search
from Evergreen import evergreen_search
from HapagLloyd import hapag_search
from HMM import hmm_search
from Maersk import maersk_search
from ONE import one_search
from ExcelFile import *
import time
from openpyxl.styles import PatternFill, Alignment 
from openpyxl import load_workbook


def modify_sheets():
    greenFill = PatternFill(start_color='8FB547', end_color='8FB547', fill_type='solid')
    redFill = PatternFill(start_color='FF7F7F', end_color='FF7F7F', fill_type='solid')
    yellowFill = PatternFill(start_color='F1EB9C', end_color='F1EB9C', fill_type='solid')
    blue_fill = PatternFill(start_color='84C4E4', end_color='84C4E4', fill_type='solid')



    cont_list = db_get_all_containers()
    excel_file_list = db_get_all_excel_files()
    #unmodified_rows = db_get_no_search_cont()

    for xcel_sheet in excel_file_list:
        workbook = load_workbook(filename=xcel_sheet[0])
        sheet = workbook[xcel_sheet[1]]
        for cont in cont_list:
            if sheet.title in cont.wb_sheet:
                cont_sheet_index = cont.wb_sheet.index(sheet.title)
                date_cell_loc = cont.date_cell_location[cont_sheet_index]
                if cont.arrival_date == 'arrived':
                    sheet[date_cell_loc] = 'arrived'
                    sheet[date_cell_loc].fill = greenFill
                    sheet[date_cell_loc].alignment = Alignment(horizontal='right')
                elif cont.arrival_date == 'Date Error':
                    sheet[date_cell_loc] = 'Date Error'
                    sheet[date_cell_loc].fill = yellowFill
                    sheet[date_cell_loc].alignment = Alignment(horizontal='right')
                else:
                    sheet[date_cell_loc] = cont.arrival_date
                    sheet[date_cell_loc].fill = blue_fill
                    sheet[date_cell_loc].alignment = Alignment(horizontal='right')

        workbook.save(xcel_sheet[0])


def main():
    #prompt to update uc chrome and regular chrome if needed, maybe do manually?

    xls_1 = ExcelFile(r'C:\Users\jmattison\Desktop\ecopax-shipping-updater\Shipping Excel Sheet Test.xlsx')
    xls_1.parse_workbook()

    xls_2 = ExcelFile(r'C:\Users\jmattison\Desktop\Container Nums\POManagementReport - 2017.xlsx')
    xls_2.parse_workbook()

    xls_3 = ExcelFile(r'C:\Users\jmattison\Desktop\Container Nums\POManagementReport - 2018.xlsx')
    xls_3.parse_workbook()

    xls_4 = ExcelFile(r'C:\Users\jmattison\Desktop\Container Nums\POManagementReport - 2020.xlsx')
    xls_4.parse_workbook()

    xls_5 = ExcelFile(r'C:\Users\jmattison\Desktop\Container Nums\POManagementReport - 2021.xlsx')
    xls_5.parse_workbook()

    xls_6 = ExcelFile(r'C:\Users\jmattison\Desktop\Container Nums\POManagementReport - 2022.xlsx')
    xls_6.parse_workbook()
    

#py2exe for executable 

if __name__ == '__main__': 
    '''
    gui_frame = ShippingUpdaterGUI()
    gui_frame.run_gui()
    
    '''
    main()
    '''
    cosco_search()
    
    evergreen_search()
    
    hmm_search()
    
    maersk_search()
    
    one_search()
    
    cma_cont_list = db_get_containers_by_carrier('CMA CGM')
    cma_search(cma_cont_list)
    
    hapag_search()
    
    '''


    
    start_time = time.perf_counter()
    
    p1 = multiprocessing.Process(target=cosco_search)
    p2 = multiprocessing.Process(target=evergreen_search)
    p4 = multiprocessing.Process(target=maersk_search)
    p5 = multiprocessing.Process(target=one_search)

    cma_list = get_divided_containers_by_carrier('CMA CGM')
    cma_process_lst = []

    for lst_chunk in cma_list:
        p_cma = multiprocessing.Process(target=cma_search, args=(lst_chunk,))
        cma_process_lst.append(p_cma)

    p1.start()
    p2.start()
    p4.start()
    p5.start()

    for srch_process in cma_process_lst:
        srch_process.start()

    p1.join()
    p2.join()
    p4.join()
    p5.join()

    for srch_process in cma_process_lst:
        srch_process.join()

    hmm_search()
    hapag_search()
    


    print(f'\n\nDone, Time Ran: {(time.perf_counter() - start_time)/60} minutes')
    
    modify_sheets()



    i = 1
   

