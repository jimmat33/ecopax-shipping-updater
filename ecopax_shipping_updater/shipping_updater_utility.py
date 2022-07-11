'''
docstr
'''
import re
import time
import random
import numpy as np
from openpyxl.styles import PatternFill, Alignment 
from openpyxl import load_workbook
from shipping_container_db import db_get_cont_by_carrier


def random_sleep():
    '''
    docstr
    '''
    sleep_time = random.uniform(0.25, 1.25)
    time.sleep(sleep_time)


def get_month_num(month):
    '''
    This function takes a month as a word and returns it as the respective number of the month for
    proper date formatting
    '''

    if_dict = {
        '01': ['January', 'JAN', 'Jan'],
        '02': ['February', 'FEB', 'Feb'],
        '03': ['March', 'MAR', 'Mar'],
        '04': ['April', 'APR', 'Apr'],
        '05': ['May', 'MAY', 'May'],
        '06': ['June', 'JUN', 'Jun'],
        '07': ['July', 'JUL', 'Jul'],
        '08': ['August', 'AUG', 'Aug'],
        '09': ['September', 'SEP', 'Sep'],
        '10': ['October', 'OCT', 'Oct'],
        '11': ['November', 'NOV', 'Nov'],
        '12': ['December', 'DEC', 'Dec']
  }

    for key in if_dict.items():
        if month in if_dict[key]:
            return key

    return 'ERROR'

def get_date_from_cma(given_str):
    '''
    This function is specifically used for CMA CGM, it takes the raw data from their website and
    returns a string that can be used to create the proper date for adding to the container's
    dictionary entry
    '''
    start_index = 0

    # This if block is checking for the days of the week and then using that as a basis for the
    # starting index of the date
    if given_str.find("Sunday") != -1:
        start_index = given_str.find("Sunday") + len("Sunday") + 1
    elif given_str.find("Monday") != -1:
        start_index = given_str.find("Monday") + len("Monday") + 1
    elif given_str.find("Tuesday") != -1:
        start_index = given_str.find("Tuesday") + len("Tuesday") + 1 
    elif given_str.find("Wednesday") != -1:
        start_index = given_str.find("Wednesday") + len("Wednesday") + 1
    elif given_str.find("Thursday") != -1:
        start_index = given_str.find("Thursday") + len("Thursday") + 1
    elif given_str.find("Friday") != -1:
        start_index = given_str.find("Friday") + len("Friday") + 1
    elif given_str.find("Saturday") != -1:
        start_index = given_str.find("Saturday") + len("Saturday") + 1
    elif given_str.find("Sunday") != -1:
        start_index = given_str.find("Sunday") + len("Sunday") + 1
    else:
        return 'ERROR'

    actual_date = given_str[start_index:(start_index + 11)]

    return actual_date


def get_divided_containers_by_carrier(carrier_company):
    '''
    docstr
    '''
    total_list = db_get_cont_by_carrier(carrier_company)
    ret_list = np.array_split(total_list, 3)

    return ret_list


def modify_sheets():
    '''
    docstr
    '''
    green_fill = PatternFill(start_color='8FB547', end_color='8FB547', fill_type='solid')
    red_fill = PatternFill(start_color='FF7F7F', end_color='FF7F7F', fill_type='solid')
    yellow_fill = PatternFill(start_color='F1EB9C', end_color='F1EB9C', fill_type='solid')
    blue_fill = PatternFill(start_color='84C4E4', end_color='84C4E4', fill_type='solid')
    no_fill = PatternFill(fill_type=None)

    cont_list = db_get_all_containers()
    no_search_list = db_get_no_search_cont()
    excel_file_list = db_get_all_excel_files()
    unmod_cont_list = db_get_all_unmod_containers()

    for xcel_sheet in excel_file_list:

        workbook = load_workbook(filename=xcel_sheet[0])
        sheet = workbook[xcel_sheet[1]]
        if type(cont_list) == list:
            
            for cont in cont_list:

                if sheet.title in cont.wb_sheet:
                    cont_sheet_index = cont.wb_sheet.index(sheet.title)
                    date_cell_loc = cont.date_cell_location[cont_sheet_index]

                    if cont.arrival_date == 'arrived':
                        sheet[date_cell_loc] = 'arrived'
                        sheet[date_cell_loc].fill = green_fill
                        sheet[date_cell_loc].alignment = Alignment(horizontal='right')
                    elif cont.arrival_date == 'Date Error':
                        sheet[date_cell_loc] = 'Date Error'
                        sheet[date_cell_loc].fill = yellow_fill
                        sheet[date_cell_loc].alignment = Alignment(horizontal='right')
                    else:
                        sheet[date_cell_loc] = cont.arrival_date
                        sheet[date_cell_loc].fill = blue_fill
                        sheet[date_cell_loc].alignment = Alignment(horizontal='right')


            for ns_cont in no_search_list:
                if sheet.title in ns_cont.wb_sheet:
                    cont_sheet_index = ns_cont.wb_sheet.index(sheet.title)
                    cont_num_cell_loc = ns_cont.container_num_cell_location[cont_sheet_index]
                    sheet[cont_num_cell_loc].fill = red_fill


            for cont in unmod_cont_list:

                if sheet.title in cont.wb_sheet:
                    cont_sheet_index = cont.wb_sheet.index(sheet.title)
                    date_cell_loc = cont.date_cell_location[cont_sheet_index]

                    if cont.arrival_date != 'arrived':
                        sheet[date_cell_loc].fill = no_fill

        workbook.save(xcel_sheet[0])
