from datetime import datetime
import time
import pandas as pd
import undetected_chromedriver as uc
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import os
from os import listdir
from os.path import isfile, join
from openpyxl.styles import PatternFill, Alignment 
from openpyxl import load_workbook

#Global lists of all modified cells in each sheet
modified_custom_cells_list = list()
modified_rest_cells_list = list()

def get_date_from_cma(given_str):
    '''
    This function is specifically used for CMA CGM, it takes the raw data from their website and
    returns a string that can be used to create the proper date for adding to the container's
    dictionary entry
    '''
    start_index = 0

    #This if block is checking for the days of the week and then using that as a basis for the starting index of the date
    if given_str.find("Sunday") != -1:
        start_index = given_str.find("Sunday") + len("Sunday") + 1
    elif given_str.find("Monday") != -1:
        start_index = given_str.find("Monday") + len("Monday") + 1
    elif given_str.find("Tuesday") != -1:
        start_index = given_str.find("Tuesday") + len("Tuesday") + 1 
    elif given_str.find("Wednesday") != -1:
        start_index = given_str.find("Wednesday") + len("Wednesday") + 1
    elif given_str.find("Thursday") != -1:
        start_index = given_str.find("Thursday") + len("Thursday") + 1
    elif given_str.find("Friday") != -1:
        start_index = given_str.find("Friday") + len("Friday") + 1
    elif given_str.find("Saturday") != -1:
        start_index = given_str.find("Saturday") + len("Saturday") + 1
    elif given_str.find("Sunday") != -1:
        start_index = given_str.find("Sunday") + len("Sunday") + 1
    else:
        return 'ERROR'

    actual_date = given_str[start_index:(start_index + 11)]

    return actual_date
    
def get_month_num(month):
    '''
    This function takes a month as a word and returns it as the respective number of the month for
    proper date formatting
    '''
    if month == 'January' or month == 'JAN' or month == 'Jan':
        return '01'
    elif month == 'February' or month == 'FEB' or month == 'Feb':
        return '02'
    elif month == 'March' or month == 'MAR' or month == 'Mar':
        return '03'
    elif month == 'April' or month == 'APR' or month == 'Apr':
        return '04'
    elif month == 'May' or month == 'MAY' or month == 'May':
        return '05'
    elif month == 'June' or month == 'JUN' or month == 'Jun':
        return '06'
    elif month == 'July' or month == 'JUL' or month == 'Jul':
        return '07'
    elif month == 'August' or month == 'AUG' or month == 'Aug':
        return '08'
    elif month == 'September' or month == 'SEP' or month == 'Sep':
        return '09'
    elif month == 'October' or month == 'OCT' or month == 'Oct':
        return '10'
    elif month == 'November' or month == 'NOV' or month == 'Nov':
        return '11'
    elif month == 'December' or month == 'DEC' or month == 'Dec':
        return '12'
    else:
        return 'ERROR'

def replace_values(date_dict,df, sheet_name):
    dict_keys = [*date_dict]
    workbook = load_workbook(filename=r'C:\Users\jmattison\Desktop\ecopax-shipping-updater\Shipping Excel Sheet.xlsx')
    sheet = workbook.get_sheet_by_name(sheet_name)

    greenFill = PatternFill(start_color='00FF00',
                   end_color='00FF00',
                   fill_type='solid')


    for key in dict_keys:
        container_num = key
        container_arrival_date = date_dict[key]

        formatted_date = container_arrival_date[2] + '-' + container_arrival_date[0] + '-' + container_arrival_date[1]
        index_list = df[df == container_num].stack().index.tolist()
        excel_date = container_arrival_date[0] + '/' + container_arrival_date[1] + '/' + container_arrival_date[2]

        if sheet_name == 'custom':
            old_date = str(df.iat[index_list[0][0], 6])

            old_formatted_date = old_date[0:10]

            new_date_datetime = datetime.strptime(formatted_date, "%Y-%m-%d")
            today_date = datetime.today()

            if old_date == 'arrived':
                print('already arrived')
            elif old_formatted_date == formatted_date or new_date_datetime < today_date:
                sheet_index = 'G' + str(index_list[0][0] + 2)
                sheet[sheet_index] = 'arrived'
                sheet[sheet_index].fill = greenFill
                sheet[sheet_index].alignment = Alignment(horizontal='right')
                
                modified_custom_cells_list.append(container_num)
            else:
                sheet_index = 'G' + str(index_list[0][0] + 2)
                sheet[sheet_index] = excel_date
                sheet[sheet_index].fill = greenFill
                sheet[sheet_index].alignment = Alignment(horizontal='right')

                modified_custom_cells_list.append(container_num)
        else:
            old_date = str(df.iat[index_list[0][0], 7])
            old_formatted_date = old_date[0:10]

            if old_date == 'arrived':
                print('already arrived')
            elif old_formatted_date == formatted_date:
                sheet_index = 'H' + str(index_list[0][0] + 2)
                sheet[sheet_index] = 'arrived'
                sheet[sheet_index].fill = greenFill
                sheet[sheet_index].alignment = Alignment(horizontal='right')

                modified_rest_cells_list.append(container_num)
            else:
                sheet_index = 'H' + str(index_list[0][0] + 2)
                sheet[sheet_index] = excel_date
                sheet[sheet_index].fill = greenFill
                sheet[sheet_index].alignment = Alignment(horizontal='right')

                modified_rest_cells_list.append(container_num)


        workbook.save(filename=r'C:\Users\jmattison\Desktop\ecopax-shipping-updater\Shipping Excel Sheet.xlsx')

def cosco_search(container_num_list):
    '''
    This function searches the cosco site for the estimated arrival date of a list of crate numbers
    '''
    return_dict = dict() 
    try:
        try:
            #setting up driver options
            options = webdriver.ChromeOptions()
            options.add_experimental_option('excludeSwitches', ['enable-logging'])
            options.add_argument('--headless')
            options.add_argument('--disable-gpu')

            #This is setting up the initial driver connection
            driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
            cosco_web_link = 'https://elines.coscoshipping.com/ebusiness/cargoTracking?trackingType=CONTAINER&number='
            cosco_search_link = cosco_web_link + container_num_list[0]
            driver.implicitly_wait(0.5)
            driver.get(cosco_search_link)

        except Exception:
            print('----------------------------------------------------------------------------------------------')
            print('-------------------------------- Cosco Site Connection Failed --------------------------------')
            print('----------------------------------------------------------------------------------------------')


        print('\n[Connection Alert] Driver Connection to Cosco Site Successful\n')


        try:
            #clicking past the TOS popup
            driver.find_element(By.XPATH, '/html/body/div[3]/div[2]/div/div/div[3]/div/button').click()

        except Exception:
            print('----------------------------------------------------------------------------------------------')
            print('------------------------------- Failed to click past Cosco TOS -------------------------------')
            print('----------------------------------------------------------------------------------------------')


        try:
            #finding estimated arrival date in website table
            web_date = driver.find_element(By.XPATH,'/html/body/div[1]/div[4]/div[1]/div/div[2]/div/div/div[2]/div[1]/div/div/div[1]/div[2]/p[2]')
        
            #waiting to get date from page and pulling data
            time.sleep(3) 
            str_date = web_date.get_attribute('textContent')

            #properly formatting date of first index in list
            year = str_date[0:4]
            month = str_date[5:7]
            day = str_date[8:10]

            #adding date to storage data structure
            return_dict[container_num_list[0]] = [month, day, year]

        except Exception:
            print('----------------------------------------------------------------------------------------------')
            print('----------------------------- Failed to find/process date Cosco ------------------------------')
            print('----------------------------------------------------------------------------------------------')


        i = 1

        while i < len(container_num_list):
            try:
                #clearning textbox, entering in new container, and clicking search button
                textbox = driver.find_element(By.XPATH, '/html/body/div[1]/div[4]/div[1]/div/div[1]/div/div[2]/form/div/div[1]/div/div/div/input')
                textbox.clear()
                textbox.send_keys(container_num_list[i])
                driver.find_element(By.XPATH, '/html/body/div[1]/div[4]/div[1]/div/div[1]/div/div[2]/form/div/div[2]/button').click()

                time.sleep(3)

                #pulling date from cosco table
                str_date = WebDriverWait(driver, 20).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[1]/div[4]/div[1]/div/div[2]/div/div/div[2]/div[1]/div/div/div[1]/div[2]/p[2]'))).get_attribute('textContent')

                #formatting date properly
                year = str_date[0:4]
                month = str_date[5:7]
                day = str_date[8:10]

                #adding date to storage data structure
                return_dict[container_num_list[i]] = [month, day, year]

            except Exception:
                print('----------------------------------------------------------------------------------------------')
                print('----------------------------- Failed to find/process date Cosco ------------------------------')
                print('----------------------------------------------------------------------------------------------')

            i += 1

        driver.close()

    except Exception:
        print('----------------------------------------------------------------------------------------------')
        print('-------------------------------- Cosco Search Major Failure ----------------------------------')
        print('----------------------------------------------------------------------------------------------')


    return return_dict


def one_search(container_num_list):
    '''
    This function searches the one site for the estimated arrival date of a list of crate numbers
    '''
    return_dict = dict()
    try:
        try:
            #Setting up driver options
            options = webdriver.ChromeOptions()
            options.add_experimental_option('excludeSwitches', ['enable-logging'])
            options.add_argument('--headless')
            options.add_argument('--disable-gpu')

            #Performing site connection
            driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
            one_link = 'https://ecomm.one-line.com/one-ecom/manage-shipment/cargo-tracking'
            driver.implicitly_wait(0.5)
            driver.get(one_link)

        except Exception:
            print('----------------------------------------------------------------------------------------------')
            print('--------------------------------- ONE Site Connection Failed ---------------------------------')
            print('----------------------------------------------------------------------------------------------')


        print('\n[Connection Alert] Driver Connection to ONE Site Successful\n')


        try:
            #Finding the frame so site can be interactive
            frame = driver.find_element(By.CSS_SELECTOR, '#__next > main > div > div.MainLayout_one-container__4HS_a > div > div.IframeCurrentEcom_wrapper__qvnCe > iframe')
            driver.switch_to.frame(frame)

            #Selecting container number from dropdown
            select = Select(driver.find_element(By.NAME, 'searchType'))
            select.select_by_visible_text('Container No.')

            #Entering container number in textbox and clicking for search
            textbox = driver.find_element(By.XPATH, '/html/body/div[3]/div[2]/div[1]/form/table/tbody/tr/td/div/textarea')
            textbox.clear()
            textbox.send_keys(container_num_list[0])
            driver.find_element(By.XPATH, '/html/body/div[3]/div[2]/div[1]/form/div[1]/div[1]/button/span').click()

            time.sleep(0.5)

        except:
            print('----------------------------------------------------------------------------------------------')
            print('------------------------------- Failed to use ONE Site Search --------------------------------')
            print('----------------------------------------------------------------------------------------------')


        try:
            #Pulling Date from site
            arrival_text = driver.find_element(By.XPATH, '/html/body/div[3]/div[2]/div[1]/form/div[8]/table/tbody/tr[10]/td[4]').get_attribute('textContent')

            #date formatting
            month = arrival_text[-11:-9]
            day = arrival_text[-8:-6]
            year = arrival_text[-16:-12]

            #adding date to storage data structure
            return_dict[container_num_list[0]] = [month, day, year]

        except Exception:
            print('----------------------------------------------------------------------------------------------')
            print('------------------------------ Failed to find/process date ONE -------------------------------')
            print('----------------------------------------------------------------------------------------------')


        i = 1

        while i < len(container_num_list):
            try:
                #Entering container num into the textbox and searching
                textbox = driver.find_element(By.XPATH, '/html/body/div[3]/div[2]/div[1]/form/table/tbody/tr/td/div/textarea')
                textbox.clear()
                textbox.send_keys(container_num_list[i])
                driver.find_element(By.XPATH, '/html/body/div[3]/div[2]/div[1]/form/div[1]/div[1]/button/span').click()

                time.sleep(0.5)

                #pulling date from site
                arrival_text = driver.find_element(By.XPATH, '/html/body/div[3]/div[2]/div[1]/form/div[8]/table/tbody/tr[10]/td[4]').get_attribute('textContent')

                #formatting date
                month = arrival_text[-11:-9]
                day = arrival_text[-8:-6]
                year = arrival_text[-16:-12]

                #adding date to storage data structure
                return_dict[container_num_list[i]] = [month, day, year]

            except Exception:
                print('----------------------------------------------------------------------------------------------')
                print('------------------------------ Failed to find/process date ONE -------------------------------')
                print('----------------------------------------------------------------------------------------------')


            i += 1

        driver.close()

    except Exception: 
        print('----------------------------------------------------------------------------------------------')
        print('--------------------------------- ONE Search Major Failure -----------------------------------')
        print('----------------------------------------------------------------------------------------------')


    return return_dict 


def hapag_search(container_num_list):
    '''
    This function searches the hapag-lloyd site for the estimated arrival date of a crate
    '''
    return_dict = dict()

    driver = uc.Chrome()
    hapag_link = 'https://www.hapag-lloyd.com/en/online-business/track/track-by-container-solution.html'
    driver.implicitly_wait(0.5)
    driver.get(hapag_link)

    try:
        WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[11]/div[2]/div[3]/div[1]/button[2]'))).click()
    except:
        driver.find_element(By.ID, "accept-recommended-btn-handler").click()

    time.sleep(0.5)

    textbox = driver.find_element_by_xpath('/html/body/div[5]/div[2]/div[2]/div/div/div/div/div[2]/div/div/div[1]/div/form/div[4]/div[2]/div/div/table/tbody/tr/td/table/tbody/tr[3]/td/table/tbody/tr/td/div/table/tbody/tr/td[1]/table/tbody/tr/td/table/tbody/tr/td[1]/table/tbody/tr/td[2]/input')
    textbox.clear()
    textbox.send_keys(container_num_list[0])

    driver.find_element_by_xpath('/html/body/div[5]/div[2]/div[2]/div/div/div/div/div[2]/div/div/div[1]/div/form/div[4]/div[2]/div/div/div[1]/table/tbody/tr/td[1]/button').click()
    
    time.sleep(0.5)
    
    table_text = driver.find_element_by_xpath('/html/body/div[5]/div[2]/div[2]/div/div/div/div/div[2]/div/div/div[1]/div/form/div[5]/div[2]/div/div/table/tbody/tr/td/table/tbody/tr[5]/td/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[2]/table')
    str_text = table_text.get_attribute('textContent')
    index = str_text.find("Vessel arrival")

    target_text = str_text[269:]
    date_text = target_text[(target_text.find("20")):]

    year = date_text[0:4]
    month = date_text[5:7]
    day = date_text[8:10]

    return_dict[container_num_list[0]] = [month, day, year]

    i = 1

    while i < len(container_num_list):
        textbox = driver.find_element_by_xpath('/html/body/div[5]/div[2]/div[2]/div/div/div/div/div[2]/div/div/div[1]/div/form/div[4]/div[2]/div/div/table/tbody/tr/td/table/tbody/tr[3]/td/table/tbody/tr/td/div/table/tbody/tr/td[1]/table/tbody/tr/td/table/tbody/tr/td[1]/table/tbody/tr/td[2]/input')
        textbox.clear()

        textbox.send_keys(container_num_list[i])
        driver.find_element_by_xpath('/html/body/div[5]/div[2]/div[2]/div/div/div/div/div[2]/div/div/div[1]/div/form/div[4]/div[2]/div/div/div[1]/table/tbody/tr/td[1]/button').click()
        time.sleep(0.5)

        time.sleep(0.5)
    
        table_text = driver.find_element_by_xpath('/html/body/div[5]/div[2]/div[2]/div/div/div/div/div[2]/div/div/div[1]/div/form/div[5]/div[2]/div/div/table/tbody/tr/td/table/tbody/tr[5]/td/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[2]/table')
        str_text = table_text.get_attribute('textContent')
        index = str_text.find("Vessel arrival")

        target_text = str_text[269:]
        date_text = target_text[(target_text.find("20")):]

        year = date_text[0:4]
        month = date_text[5:7]
        day = date_text[8:10]

        return_dict[container_num_list[i]] = [month, day, year]

        i += 1

    driver.close()
    return return_dict

def maersk_search(container_num):
    '''
    This function searches the maersk site for the estimated arrival date of a crate
    '''
    maersk_link = 'https://www.maersk.com/tracking/'
    driver = webdriver.Chrome(executable_path=r'C:\Users\jmattison\Desktop\ecopax-shipping-updater\chromedriver.exe')

    maersk_link = maersk_link + container_num

    driver.implicitly_wait(0.5)
    driver.get(maersk_link)

    try:
        WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div/div/div[1]/div[2]/button[3]'))).click()
    except:
        driver.find_elements_by_class_name('coi-banner__accept--fixed-margin').click()

    str_text = WebDriverWait(driver, 20).until(EC.visibility_of_element_located((By.XPATH, '/html/body/main/div/div/div[3]/dl/dd[1]'))).get_attribute('textContent')

    year = str(str_text[-5:-1].strip())

    month = str((str_text[3:-5]).strip())
    month_num = get_month_num(month)

    day = str(str_text[0:3].strip())

    return [month_num, day, year]
   
def cma_search(container_num_list):
    return_dict = dict()

    cma_link = 'https://www.cma-cgm.com/ebusiness/tracking'
    speech_to_text_link = 'https://speech-to-text-demo.ng.bluemix.net/'
    options = webdriver.ChromeOptions()
    options.add_argument('--window-size=1100,1000')
    
    prefs = {"profile.default_content_settings.popups": 0, "download.default_directory": r"C:\Users\jmattison\Desktop\ecopax-shipping-updater\Audio Captcha Files\\", "directory_upgrade": True}
    options.add_experimental_option("prefs", prefs)

    driver = webdriver.Chrome(executable_path=r'C:\Users\jmattison\Desktop\ecopax-shipping-updater\chromedriver.exe', options=options)
    driver.get(cma_link)

    #---------------------------------------------------------------------------------------------------------------------------------------------------
    #---------------------------------------------------------------------------------------------------------------------------------------------------
    #------------------------------------Audio Bypass---------------------------------------------------------------------------------------------------
    frame = driver.find_element_by_css_selector('body > iframe')
    driver.switch_to.frame(frame)

    try:
        WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div/div[3]/div/div[3]/div/div/div[2]/div[1]'))).click()
    except Exception:
        print('error')

    try:
        WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[2]/div[1]/div/div[2]/div/a[4]'))).click()
    except Exception:
        print('error')

    time.sleep(1)
    audio_src_link = WebDriverWait(driver, 10).until(EC.invisibility_of_element((By.TAG_NAME, 'audio'))).get_attribute('currentSrc')

    driver.execute_script("window.open('');")
  
    # Switch to the new window and open new URL
    driver.switch_to.window(driver.window_handles[1])
    driver.get(audio_src_link)

    driver.execute_script('''let aLink = document.createElement("a");let videoSrc = document.querySelector("video").firstChild.src;aLink.href = videoSrc;aLink.download = "";aLink.click();aLink.remove();''')


    driver.execute_script("window.open('');")
  
    # Switch to the new window and open new URL
    driver.switch_to.window(driver.window_handles[2])
    driver.get(speech_to_text_link)
    
    mypath = r'C:\Users\jmattison\Desktop\ecopax-shipping-updater\Audio Captcha Files'

    onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]

    root = driver.find_element_by_id('root').find_elements_by_class_name('dropzone _container _container_large')
    btn = driver.find_element(By.XPATH, '//*[@id="root"]/div/input')

    file_str = onlyfiles[-1]
    send_keys_str = r'C:\Users\jmattison\Desktop\ecopax-shipping-updater\Audio Captcha Files' + '\\' + file_str

    time.sleep(5)
    btn.send_keys(send_keys_str)

    time.sleep(15)
    #btn.send_keys(path)

    # Audio to text is processing
    text = driver.find_element(By.XPATH, '//*[@id="root"]/div/div[7]/div/div/div').find_elements_by_tag_name('span')

    result = " ".join( [ each.text for each in text ] )
    key_str = result[33:-1]

    driver.switch_to.window(driver.window_handles[0])
    frame = driver.find_element_by_css_selector('body > iframe')
    driver.switch_to.frame(frame)

    textbox = driver.find_element_by_xpath('/html/body/div[2]/div[2]/div[1]/div/div/div/div[2]/div[3]/input')
    textbox.send_keys(key_str)

    driver.find_element_by_xpath('/html/body/div[2]/div[2]/div[1]/div/div/div/div[2]/div[4]').click()

    os.remove(send_keys_str)
    #------------------------------------Audio Bypass---------------------------------------------------------------------------------------------------
    #---------------------------------------------------------------------------------------------------------------------------------------------------
    #---------------------------------------------------------------------------------------------------------------------------------------------------

    try:
        WebDriverWait(driver, 20).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[3]/main/section/div/div[2]/fieldset/form[3]/div/span[1]/input[2]')))
    except Exception:
        print('error')
    
    textbox = driver.find_element_by_xpath('/html/body/div[3]/main/section/div/div[2]/fieldset/form[3]/div/span[1]/input[2]')
    textbox.send_keys(container_num_list[0])

    driver.find_element_by_xpath('/html/body/div[3]/main/section/div/div[2]/fieldset/form[3]/p/button').click()

    time.sleep(0.5)
    try:
        str_date = driver.find_element_by_css_selector('#trackingsearchsection > div > section > div > div > div').get_attribute('textContent')
        useable_date = get_date_from_cma(str_date)

        month = get_month_num(useable_date[3:6])
        day = useable_date[0:2]
        year = useable_date[7:]

        return_dict[container_num_list[0]] = [month, day, year]
    except Exception:
        print('error')

    

    i = 1

    while i < len(container_num_list):
        try:
            WebDriverWait(driver, 20).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[3]/main/section/div/div[2]/fieldset/form[3]/div/span[1]/input[2]')))
        except Exception:
            print('error')
    
        textbox = driver.find_element_by_xpath('/html/body/div[3]/main/section/div/div[2]/fieldset/form[3]/div/span[1]/input[2]')
        textbox.clear()
        textbox.send_keys(container_num_list[i])

        driver.find_element_by_xpath('/html/body/div[3]/main/section/div/div[2]/fieldset/form[3]/p/button').click()

        time.sleep(0.5)

        str_date = driver.find_element_by_css_selector('#trackingsearchsection > div > section > div > div > div').get_attribute('textContent')
        useable_date = get_date_from_cma(str_date)

        month = get_month_num(useable_date[3:6])
        day = useable_date[0:2]
        year = useable_date[7:]

        return_dict[container_num_list[i]] = [month, day, year]

        i += 1

    driver.close()
    return return_dict

def evergreen_search(container_num_list):
    return_dict = dict()

    evergreen_link = 'https://ct.shipmentlink.com/servlet/TDB1_CargoTracking.do'
    driver = webdriver.Chrome(executable_path=r'C:\Users\jmattison\Desktop\ecopax-shipping-updater\chromedriver.exe')
    driver.implicitly_wait(0.5)
    driver.get(evergreen_link)

    driver.find_element_by_xpath('/html/body/div[4]/center/table[2]/tbody/tr/td/form/span[2]/table[2]/tbody/tr[1]/td/table/tbody/tr/td[1]/table/tbody/tr[2]/td[2]/input').click()

    textbox = driver.find_element_by_xpath('/html/body/div[4]/center/table[2]/tbody/tr/td/form/span[2]/table[2]/tbody/tr[1]/td/table/tbody/tr/td[2]/input[1]')
    textbox.send_keys(container_num_list[0])

    driver.find_element_by_xpath('/html/body/div[4]/center/table[2]/tbody/tr/td/form/span[2]/table[2]/tbody/tr[1]/td/table/tbody/tr/td[2]/input[2]').click()

    str_date = driver.find_element_by_xpath('/html/body/div[5]/center/table[2]/tbody/tr/td/table[2]/tbody/tr/td').get_attribute('textContent')

    month = str_date[28:31]
    day = str_date[32:34]
    year = str_date[35:]

    month_num = get_month_num(month)
    
    return_dict[container_num_list[0]] = [month_num, day, year]

    i = 1

    while i < len(container_num_list):
        textbox = driver.find_element_by_xpath('/html/body/div[5]/center/table[1]/tbody/tr/td/form/table/tbody/tr/td/table/tbody/tr/td[2]/input[1]')
        textbox.clear()
        textbox.send_keys(container_num_list[i])

        driver.find_element_by_xpath('/html/body/div[5]/center/table[1]/tbody/tr/td/form/table/tbody/tr/td/table/tbody/tr/td[2]/input[2]').click()

        time.sleep(0.5)
        str_date = driver.find_element_by_xpath('/html/body/div[5]/center/table[2]/tbody/tr/td/table[2]/tbody/tr/td').get_attribute('textContent')

        month = str_date[28:31]
        day = str_date[32:34]
        year = str_date[35:]

        month_num = get_month_num(month)
    
        return_dict[container_num_list[i]] = [month_num, day, year]

        i += 1

    driver.close()
    return return_dict

def hmm_search(container_num):
    return_dict = dict()

    hmm_link_part_1 = 'https://www.hmm21.com/cms/company/engn/index.jsp?type=2&number='
    hmm_link_part_2 = '&is_quick=Y&quick_params='
    driver = uc.Chrome()
    driver.implicitly_wait(0.5)

    hmm_search_link = hmm_link_part_1 + container_num + hmm_link_part_2

    driver.get(hmm_search_link)

    selector = Select(driver.find_element_by_xpath('/html/body/div[2]/div[1]/div/div[2]/form/fieldset/p/span[1]/select'))
    selector.select_by_visible_text('Container No.')

    driver.find_element_by_xpath('/html/body/div[2]/div[1]/div/div[2]/form/fieldset/p/span[3]/a').click()

    # Store the ID of the original window
    original_window = driver.current_window_handle

    wait = WebDriverWait(driver, 3)

    wait.until(EC.number_of_windows_to_be(2))

    # Loop through until we find a new window handle
    for window_handle in driver.window_handles:
        if window_handle != original_window:
            driver.switch_to.window(window_handle)
            break

    # Wait for the new tab to finish loading content
    wait.until(EC.title_is("HMM Customer Plus"))

    try:
        WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div/div/div[3]/button[2]'))).click()
    except Exception:
        print('exception thrown')

    time.sleep(0.5)

    frame = driver.find_element_by_css_selector('#_frame1')
    driver.switch_to.frame(frame)

    str_date = WebDriverWait(driver, 20).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[4]/form[1]/div/div[6]/table/tbody/tr[3]/td[5]/span'))).get_attribute('textContent')

    month = str_date[3:6]
    day = str_date[0:2]
    year = str_date[7:11:]

    month_num = get_month_num(month)
    
    driver.close()
    return [month_num, day, year]

def main():
    #put chrome driver version check and chrome version check, prompt to update if necessary


    #Setting up all the lists for each 
    custom_cosco_list = list()
    rest_cosco_list = list()

    custom_one_list = list()
    rest_one_list = list()

    custom_hapag_list = list()
    rest_hapag_list = list()

    custom_maersk_list = list()
    rest_maersk_list = list()

    custom_cma_list = list()
    rest_cma_list = list()
    
    custom_evergreen_list = list()
    rest_evergreen_list = list()

    custom_hmm_list = list()
    rest_hmm_list = list()

    #getting data to pull from for both sheets
    customsheet_data = pd.read_excel(r'C:\Users\jmattison\Desktop\ecopax-shipping-updater\Shipping Excel Sheet.xlsx', sheet_name = 'custom')
    restsheet_data = pd.read_excel(r'C:\Users\jmattison\Desktop\ecopax-shipping-updater\Shipping Excel Sheet.xlsx', sheet_name = 'Rest')

    #populating lists for the custom sheet
    for index,row in customsheet_data.iterrows():
        if row['Carrier'] == 'Cosco':
            custom_cosco_list.append(row['Container Number'])
        elif row['Carrier'] == 'ONE':
            custom_one_list.append(row['Container Number'])
        elif row['Carrier'] == 'Hapag-Lloyd':
            custom_hapag_list.append(row['Container Number'])
        elif row['Carrier'] == 'Maersk':
            custom_maersk_list.append(row['Container Number'])
        elif row['Carrier'] == 'CMA CGM':
            custom_cma_list.append(row['Container Number'])
        elif row['Carrier'] == 'Evergreen':
            custom_evergreen_list.append(row['Container Number'])
        elif row['Carrier'] == 'HMM':
            custom_hmm_list.append(row['Container Number'])
        else:
            print()
            #highlight red for not looked at

    #populating lists for the rest sheet
    for index, row in restsheet_data.iterrows():
        if row['Carrier'] == 'Cosco':
            rest_cosco_list.append(row['Container Number'])
        elif row['Carrier'] == 'ONE':
            rest_one_list.append(row['Container Number'])
        elif row['Carrier'] == 'Hapag-Lloyd':
            rest_hapag_list.append(row['Container Number'])
        elif row['Carrier'] == 'Maersk':
            rest_maersk_list.append(row['Container Number'])
        elif row['Carrier'] == 'CMA CGM':
            rest_cma_list.append(row['Container Number'])
        elif row['Carrier'] == 'Evergreen':
            rest_evergreen_list.append(row['Container Number'])
        elif row['Carrier'] == 'HMM':
            rest_hmm_list.append(row['Container Number'])
        else:
            print()
            #highlight red for not looked at
        
    #dictionary creation to hold all dates pulled for custom sheet containers
    cosco_custom_dates_dict = dict()
    one_custom_dates_dict = dict()
    hapag_custom_dates_dict = dict()
    maersk_custom_dates_dict = dict()
    cma_custom_dates_dict = dict()  
    evergreen_custom_dates_dict = dict()  
    hmm_custom_dates_dict = dict()

    #dictionary creation to hold all dates pulled for rest sheet containers
    cosco_rest_dates_dict = dict()
    one_rest_dates_dict = dict()
    hapag_rest_dates_dict = dict()
    maersk_rest_dates_dict = dict()
    cma_rest_dates_dict = dict()
    evergreen_rest_dates_dict = dict() 
    hmm_rest_dates_dict = dict()

    #---------------------------------------- All searches for each website----------------------------------------
    '''
    #Searching for all cosco containers from the custom sheet
    if len(custom_cosco_list) != 0:
        cosco_custom_dates_dict = cosco_search(custom_cosco_list)
        if len(cosco_custom_dates_dict) == 0:
            for i in [1,2,3]:
                print('\n[Driver Alert] Trying Costco Search Again (Custom Sheet)\n')
                cosco_custom_dates_dict = cosco_search(custom_cosco_list)
                if len(cosco_custom_dates_dict) == 0:
                    break
        if len(cosco_custom_dates_dict) == 0:
            print('\n[Driver Alert] Cosco Search Fatal Error (Custom Sheet)\n')
            #highlight all cosco custom sheet items red


    #Searching for all cosco containers from the rest sheet
    if len(rest_cosco_list) != 0:
        cosco_rest_dates_dict = cosco_search(rest_cosco_list)
        if len(cosco_rest_dates_dict) == 0:
            for i in [1,2,3]:
                print('\n[Driver Alert] Trying Costco Search Again (Rest Sheet)\n')
                cosco_rest_dates_dict = cosco_search(rest_cosco_list)
                if len(cosco_rest_dates_dict) == 0:
                    break
        if len(cosco_rest_dates_dict) == 0:
            print('\n[Driver Alert] Cosco Search Fatal Error (Rest Sheet)\n')
            #highlight all cosco rest sheet items red
    '''

    #Searching for all ONE containers from the custom sheet
    if len(custom_one_list) != 0:
        one_custom_dates_dict = one_search(custom_one_list)
        if len(one_custom_dates_dict) == 0:
            for i in [1,2,3]:
                print('\n[Driver Alert] Trying ONE Search Again (Custom Sheet)\n')
                one_custom_dates_dict = one_search(custom_one_list)
                if len(one_custom_dates_dict) == 0:
                    break
        if len(one_custom_dates_dict) == 0:
            print('\n[Driver Alert] ONE Search Fatal Error (Custom Sheet)\n')
            #highlight all ONE custom sheet items red


    #Searching for all ONE containers from the rest sheet
    if len(rest_one_list) != 0:
        one_rest_dates_dict = one_search(rest_one_list)
        if len(one_rest_dates_dict) == 0:
            for i in [1,2,3]:
                print('\n[Driver Alert] Trying ONE Search Again (rest Sheet)\n')
                one_rest_dates_dict = one_search(rest_one_list)
                if len(one_rest_dates_dict) == 0:
                    break
        if len(one_rest_dates_dict) == 0:
            print('\n[Driver Alert] ONE Search Fatal Error (rest Sheet)\n')
            #highlight all ONE rest sheet items red

    
    if len(custom_hapag_list) != 0:
        hapag_custom_dates_dict = hapag_search(custom_hapag_list)

    if len(rest_hapag_list) != 0:
        hapag_rest_dates_dict = hapag_search(rest_hapag_list)
    
    for container_num in custom_maersk_list:
        maersk_custom_dates_dict[container_num] = maersk_search(container_num)

    for container_num in rest_maersk_list:
         maersk_rest_dates_dict[container_num] = maersk_search(container_num)
    
    if len(custom_cma_list) != 0:    
        cma_custom_dates_dict = cma_search(custom_cma_list)

    if len(rest_cma_list) != 0:
        cma_rest_dates_dict = cma_search(rest_cma_list)
    
    if len(custom_evergreen_list) != 0:
        evergreen_custom_dates_dict = evergreen_search(custom_evergreen_list)

    if len(rest_evergreen_list) != 0:
        evergreen_rest_dates_dict = evergreen_search(rest_evergreen_list)

    for container_num in custom_hmm_list:
        hmm_custom_dates_dict[container_num] = hmm_search(container_num)

    for container_num in rest_hmm_list:
         hmm_rest_dates_dict[container_num] = hmm_search(container_num)
    #------------------------------------------------------------------------------------------------------------

    #Creates dictionaries to hold all dates found from searches, separated by sheet
    rest_total_dict = cosco_rest_dates_dict | one_rest_dates_dict | hapag_rest_dates_dict | maersk_rest_dates_dict | cma_rest_dates_dict | evergreen_rest_dates_dict | hmm_rest_dates_dict
    custom_total_dict =  cosco_custom_dates_dict | one_custom_dates_dict | hapag_custom_dates_dict | maersk_custom_dates_dict | cma_custom_dates_dict | evergreen_custom_dates_dict | hmm_custom_dates_dict

    #Calls function to check and replace all values on each sheet
    replace_values(rest_total_dict, restsheet_data, 'Rest')
    replace_values(custom_total_dict, customsheet_data, 'custom')

if __name__ == '__main__': 
    main()


'''
Storage for possibly future used variables

custom_yangming_list = list()
rest_yangming_list = list()
custom_msc_list = list()
rest_msc_list = list()
custom_oocl_list = list()
rest_oocl_list = list()
oocl_rest_dates_dict = dict()
msc_rest_dates_dict = dict()
yangming_rest_dates_dict = dict()
oocl_custom_dates_dict = dict()
msc_custom_dates_dict = dict()
yangming_custom_dates_dict = dict(

custom_hapag_list.append('HLXU1143116')
custom_hapag_list.append('HLXU5257457')
custom_yangming_list.append('YMLU8268863')
custom_yangming_list.append('YMLU5426986')
custom_maersk_list.append('MSKU9342870')
custom_maersk_list.append('MRKU8485175')
custom_cma_list.append('CMAU0459057')
custom_cma_list.append('CMAU1999430')
custom_msc_list.append('MSCU6919130')
custom_msc_list.append('MSCU6919130')

custom_unadded_containers_list = list()
rest_unadded_containers_list = list()

def yangming_search(container_num): #has an api
    print('No yangming')

def msc_search(container_num): #has an api
    print('no msc')

def oocl_search(container_num): #has an api
    print('no oocl')

'''